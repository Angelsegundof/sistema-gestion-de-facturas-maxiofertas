import openpyxl
import re
import datetime
import json
from collections import Counter, defaultdict

# Chilean RUT validation & canonicalization (Modulo 11)
def clean_rut(raw):
    if raw is None:
        return ""
    s = str(raw).strip().upper()
    return re.sub(r'[^0-9K]', '', s)

def validate_rut(raw):
    clean = clean_rut(raw)
    if len(clean) < 2:
        return False
    body = clean[:-1]
    dv = clean[-1]
    if not body.isdigit():
        return False
    
    # Calculate Modulo 11
    total = 0
    multiplier = 2
    for digit in reversed(body):
        total += int(digit) * multiplier
        multiplier = 2 if multiplier == 7 else multiplier + 1
    
    expected_dv_num = 11 - (total % 11)
    if expected_dv_num == 11:
        expected_dv = '0'
    elif expected_dv_num == 10:
        expected_dv = 'K'
    else:
        expected_dv = str(expected_dv_num)
        
    return dv == expected_dv

def format_rut(clean):
    if not clean or len(clean) < 2:
        return clean
    body = clean[:-1]
    dv = clean[-1]
    # Format with dots and dash: XX.XXX.XXX-X
    formatted_body = ""
    for i, digit in enumerate(reversed(body)):
        if i > 0 and i % 3 == 0:
            formatted_body = "." + formatted_body
        formatted_body = digit + formatted_body
    return f"{formatted_body}-{dv}"

# 14 Real Warehouses registered in Phase 11B.1
REGISTERED_WAREHOUSES = {
    'CENTRAL': 'Santiago Central',
    'RANCAGUA': 'Bodega Rancagua',
    'CASTRO': 'Bodega Castro Chiloé',
    'CONCEPCION': 'Bodega Concepción',
    'TEMUCO': 'Bodega Temuco',
    'TALCA': 'Bodega Talca',
    'VINA': 'Bodega Viña del Mar',
    'ANTOFAGASTA': 'Bodega Antofagasta',
    'CHILLAN': 'Bodega Chillán',
    'PUERTO_MONTT': 'Bodega Puerto Montt',
    'LOS_ANGELES': 'Bodega Los Ángeles',
    'CURICO': 'Bodega Curicó',
    'VALDIVIA': 'Bodega Valdivia',
    'LA_SERENA': 'Bodega La Serena',
}

WAREHOUSE_ALIASES = {
    'santiago': 'CENTRAL',
    'stgo': 'CENTRAL',
    'central': 'CENTRAL',
    'santiago central': 'CENTRAL',
    'rancagua': 'RANCAGUA',
    'castro': 'CASTRO',
    'castro chiloe': 'CASTRO',
    'chiloe': 'CASTRO',
    'chiloé': 'CASTRO',
    'concepcion': 'CONCEPCION',
    'concepción': 'CONCEPCION',
    'conce': 'CONCEPCION',
    'temuco': 'TEMUCO',
    'talca': 'TALCA',
    'vina': 'VINA',
    'viña': 'VINA',
    'viña del mar': 'VINA',
    'vina del mar': 'VINA',
    'antofagasta': 'ANTOFAGASTA',
    'chillan': 'CHILLAN',
    'chillán': 'CHILLAN',
    'pto montt': 'PUERTO_MONTT',
    'puerto montt': 'PUERTO_MONTT',
    'los angeles': 'LOS_ANGELES',
    'los ángeles': 'LOS_ANGELES',
    'curico': 'CURICO',
    'curicó': 'CURICO',
    'valdivia': 'VALDIVIA',
    'la serena': 'LA_SERENA',
    'laserena': 'LA_SERENA',
    'serena': 'LA_SERENA',
}

# 20 Real Users registered in Phase 11B.1
PROVISIONED_USERS = {
    'sistemasecuweb@gmail.com': {'name': 'Angel Ferrer', 'role': 'ADMIN'},
    'miyelics@gmail.com': {'name': 'Miyelis Contreras', 'role': 'MANAGEMENT'},
    'keila.maxiofertas@gmail.com': {'name': 'Keyla Contreras', 'role': 'MANAGEMENT'},
    'verocars1178@gmail.com': {'name': 'Aracelis Cardenas', 'role': 'INVOICE_EXECUTOR'},
    'maxiofertasmeli@gmail.com': {'name': 'Yuliany Alecio', 'role': 'INVOICE_EXECUTOR'},
    'jenimaxiofetas@gmail.com': {'name': 'Jeni Contreras', 'role': 'INVOICE_EXECUTOR'},
    'santiago.maxiofertas@gmail.com': {'name': 'Bodega Santiago', 'role': 'WAREHOUSE_USER', 'warehouse': 'CENTRAL'},
    'bodegarancagua13@gmail.com': {'name': 'Bodega Rancagua', 'role': 'WAREHOUSE_USER', 'warehouse': 'RANCAGUA'},
    'bodegacastrochiloe@gmail.com': {'name': 'Bodega Castro Chiloe', 'role': 'WAREHOUSE_USER', 'warehouse': 'CASTRO'},
    'megaofertas379@gmail.com': {'name': 'Bodega Concepción', 'role': 'WAREHOUSE_USER', 'warehouse': 'CONCEPCION'},
    'temuco.maxiofertas@gmail.com': {'name': 'Bodega Temuco', 'role': 'WAREHOUSE_USER', 'warehouse': 'TEMUCO'},
    'bodegatalca63@gmail.com': {'name': 'Bodega Talca', 'role': 'WAREHOUSE_USER', 'warehouse': 'TALCA'},
    'bodegavina56@gmail.com': {'name': 'Bodega Viña', 'role': 'WAREHOUSE_USER', 'warehouse': 'VINA'},
    'antofagastaonline7@gmail.com': {'name': 'Bodega Antofagasta', 'role': 'WAREHOUSE_USER', 'warehouse': 'ANTOFAGASTA'},
    'bodegachillan84@gmail.com': {'name': 'Bodega Chillán', 'role': 'WAREHOUSE_USER', 'warehouse': 'CHILLAN'},
    '1523ofertas@gmail.com': {'name': 'Bodega Pto Montt', 'role': 'WAREHOUSE_USER', 'warehouse': 'PUERTO_MONTT'},
    'bodegalosangels8@gmail.com': {'name': 'Bodega Los Angeles', 'role': 'WAREHOUSE_USER', 'warehouse': 'LOS_ANGELES'},
    'vcurico0@gmail.com': {'name': 'Bodega Curicó', 'role': 'WAREHOUSE_USER', 'warehouse': 'CURICO'},
    'bodegavaldiva829@gmail.com': {'name': 'Bodega Valdivia', 'role': 'WAREHOUSE_USER', 'warehouse': 'VALDIVIA'},
    'bodegalaserena392@gmail.com': {'name': 'Bodega la Serena', 'role': 'WAREHOUSE_USER', 'warehouse': 'LA_SERENA'},
}

def parse_amount(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(round(val))
    s = str(val).strip()
    s = s.replace('$', '').replace('.', '').replace(',', '.').strip()
    try:
        f = float(s)
        return int(round(f))
    except:
        return 0

def normalize_status(raw_status):
    if not raw_status:
        return 'PENDING'
    s = str(raw_status).strip().lower()
    if any(k in s for k in ['realizada', 'facturada', 'completada', 'emitida', 'si', 'ok', 'lista', 'facturado']):
        return 'COMPLETED'
    if any(k in s for k in ['pendiente', 'pend']):
        return 'PENDING'
    if any(k in s for k in ['en proceso', 'proceso', 'tomada']):
        return 'IN_PROGRESS'
    if any(k in s for k in ['corregir', 'correccion', 'corrección', 'rechazada', 'observacion']):
        return 'NEEDS_CORRECTION'
    if any(k in s for k in ['anulada', 'anulado', 'cancelada', 'cancelado', 'no']):
        return 'CANCELLED'
    return 'UNKNOWN'

wb = openpyxl.load_workbook('historico_facturacion.xlsx', data_only=True)
sheet = wb['Respuestas de formulario 1']

print("=== AUDITORÍA EXHAUSTIVA DE 'Respuestas de formulario 1' ===")
headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
print(f"Columnas ({len(headers)}):", headers)

total_rows = 0
valid_rows = 0
empty_rows = 0
warnings_count = 0
errors_count = 0

ruts_valid = 0
ruts_invalid = 0
unique_customers = set()
customer_conflicts = 0
customer_names_by_rut = defaultdict(set)

warehouse_counts = Counter()
unrecognized_warehouses = Counter()

status_counts = Counter()
dates_list = []
total_gross_amount = 0
amount_invalid_count = 0

requester_counts = Counter()
google_drive_links = 0
other_links = 0

duplicate_exact_fingerprints = Counter()
rows_processed = []

for r in range(2, sheet.max_row + 1):
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    if not any(v is not None for v in row_vals):
        empty_rows += 1
        continue
    
    total_rows += 1
    
    marca_temporal = row_vals[0] if len(row_vals) > 0 else None
    bodega_raw = str(row_vals[1]).strip() if len(row_vals) > 1 and row_vals[1] is not None else ""
    rut_raw = str(row_vals[2]).strip() if len(row_vals) > 2 and row_vals[2] is not None else ""
    razon_social_raw = str(row_vals[3]).strip() if len(row_vals) > 3 and row_vals[3] is not None else ""
    giro_raw = str(row_vals[4]).strip() if len(row_vals) > 4 and row_vals[4] is not None else ""
    telefono_raw = row_vals[5] if len(row_vals) > 5 else None
    correo_raw = str(row_vals[6]).strip() if len(row_vals) > 6 and row_vals[6] is not None else ""
    productos_raw = str(row_vals[7]).strip() if len(row_vals) > 7 and row_vals[7] is not None else ""
    precio_raw = row_vals[8] if len(row_vals) > 8 else None
    total_facturar_raw = row_vals[11] if len(row_vals) > 11 else None
    solicitante_raw = str(row_vals[12]).strip() if len(row_vals) > 12 and row_vals[12] is not None else ""
    estatus_raw = str(row_vals[14]).strip() if len(row_vals) > 14 and row_vals[14] is not None else ""

    # 1. Date check
    dt = None
    if isinstance(marca_temporal, datetime.datetime):
        dt = marca_temporal
        dates_list.append(dt)
    elif isinstance(marca_temporal, datetime.date):
        dt = datetime.datetime.combine(marca_temporal, datetime.time.min)
        dates_list.append(dt)
        
    # 2. Warehouse check
    b_norm = bodega_raw.lower().strip()
    mapped_wh = WAREHOUSE_ALIASES.get(b_norm)
    if mapped_wh:
        warehouse_counts[REGISTERED_WAREHOUSES[mapped_wh]] += 1
    else:
        unrecognized_warehouses[bodega_raw] += 1
        
    # 3. RUT check
    canonical_rut = clean_rut(rut_raw)
    is_valid_rut = validate_rut(rut_raw)
    if is_valid_rut:
        ruts_valid += 1
        unique_customers.add(canonical_rut)
        if razon_social_raw:
            customer_names_by_rut[canonical_rut].add(razon_social_raw.upper())
    else:
        ruts_invalid += 1
        
    # 4. Amount check
    amount = parse_amount(total_facturar_raw)
    if amount == 0 and precio_raw is not None:
        amount = parse_amount(precio_raw)
    if amount > 0:
        total_gross_amount += amount
    else:
        amount_invalid_count += 1
        
    # 5. Status check
    mapped_status = normalize_status(estatus_raw)
    status_counts[mapped_status] += 1
    
    # 6. Solicitante / Requester check
    if solicitante_raw:
        requester_counts[solicitante_raw] += 1
        
    # 7. Document link search in row values
    row_text = " ".join([str(v) for v in row_vals if v is not None])
    if 'drive.google.com' in row_text or 'docs.google.com' in row_text:
        google_drive_links += 1
    elif 'http://' in row_text or 'https://' in row_text:
        other_links += 1
        
    # 8. Duplicate fingerprint: (canonical_rut, mapped_wh, amount, dt.strftime('%Y-%m-%d') if dt else '')
    dt_str = dt.strftime('%Y-%m-%d') if dt else ''
    fingerprint = f"{canonical_rut}|{mapped_wh}|{amount}|{dt_str}"
    duplicate_exact_fingerprints[fingerprint] += 1

print(f"\n--- RESUMEN GENERAL ---")
print(f"Filas no vacías analizadas: {total_rows}")
print(f"Filas vacías omitidas: {empty_rows}")
print(f"Rango de fechas: desde {min(dates_list)} hasta {max(dates_list)}" if dates_list else "Sin fechas")
print(f"Monto total bruto CLP: ${total_gross_amount:,} CLP")
print(f"Montos válidos > 0: {total_rows - amount_invalid_count} | Montos $0/vacíos: {amount_invalid_count}")

print(f"\n--- RUTS Y CLIENTES ---")
print(f"RUTs válidos: {ruts_valid} ({ruts_valid/total_rows*100:.1f}%)")
print(f"RUTs inválidos o de prueba: {ruts_invalid}")
print(f"Clientes únicos (por RUT canónico): {len(unique_customers)}")

# Check conflicts in customer legal names
for rut, names in customer_names_by_rut.items():
    if len(names) > 1:
        customer_conflicts += 1
print(f"Clientes con múltiples variaciones de Razón Social: {customer_conflicts}")

print(f"\n--- ESTADOS DETECTADOS ---")
for st, cnt in status_counts.most_common():
    print(f"  {st:<18}: {cnt:>6} ({cnt/total_rows*100:.1f}%)")

print(f"\n--- DISTRIBUCIÓN POR BODEGA ---")
for wh, cnt in warehouse_counts.most_common():
    print(f"  {wh:<25}: {cnt:>6} ({cnt/total_rows*100:.1f}%)")

if unrecognized_warehouses:
    print(f"\nBodegas no reconocidas: {dict(unrecognized_warehouses)}")

print(f"\n--- SOLICITANTES DETECTADOS (Top 10) ---")
for req, cnt in requester_counts.most_common(10):
    print(f"  {req:<25}: {cnt:>6}")

print(f"\n--- DOCUMENTOS / ENLACES ---")
print(f"Enlaces Google Drive detectados: {google_drive_links}")
print(f"Otros enlaces: {other_links}")

exact_dups = sum(cnt - 1 for cnt in duplicate_exact_fingerprints.values() if cnt > 1)
print(f"\n--- DUPLICADOS DETECTADOS ---")
print(f"Registros con coincidencia exacta (RUT + Bodega + Monto + Mismo Día): {exact_dups}")
