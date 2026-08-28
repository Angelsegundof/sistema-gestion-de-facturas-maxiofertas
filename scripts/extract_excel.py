import openpyxl
import json
import datetime
import os
import sys

def extract(xlsx_path, json_out_path):
    if not os.path.exists(xlsx_path):
        print(f"Error: {xlsx_path} does not exist", file=sys.stderr)
        sys.exit(1)
        
    os.makedirs(os.path.dirname(os.path.abspath(json_out_path)), exist_ok=True)
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb['Respuestas de formulario 1']
    rows = []
    
    for r in range(2, sheet.max_row + 1):
        vals = [sheet.cell(r, c).value for c in range(1, 20)]
        if not any(v is not None for v in vals):
            continue
        
        dt_val = vals[0]
        dt_str = dt_val.isoformat() if isinstance(dt_val, (datetime.datetime, datetime.date)) else str(dt_val) if dt_val else None
        
        rows.append({
            'rowNum': r,
            'timestamp': dt_str,
            'bodega': str(vals[1]).strip() if vals[1] is not None else '',
            'rut': str(vals[2]).strip() if vals[2] is not None else '',
            'legalName': str(vals[3]).strip() if vals[3] is not None else '',
            'businessActivity': str(vals[4]).strip() if vals[4] is not None else '',
            'phone': str(vals[5]).strip() if vals[5] is not None else '',
            'email': str(vals[6]).strip() if vals[6] is not None else '',
            'product': str(vals[7]).strip() if vals[7] is not None else '',
            'price': vals[8],
            'totalGross': vals[11],
            'requester': str(vals[12]).strip() if vals[12] is not None else '',
            'status': str(vals[14]).strip() if vals[14] is not None else '',
            'documentUrl': str(vals[15]).strip() if vals[15] is not None else '',
            'whatsappMsg': str(vals[16]).strip() if vals[16] is not None else '',
            'notes': str(vals[17]).strip() if vals[17] is not None else '',
        })

    with open(json_out_path, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False)
    print(f"Extracted {len(rows)} rows to {json_out_path}")

if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else 'historico_facturacion.xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else '.data/temp_historical_rows.json'
    extract(xlsx, out)
