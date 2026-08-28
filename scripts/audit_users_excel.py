import openpyxl
import re
import json

wb = openpyxl.load_workbook('Usuarios sistemas de facturacion. .xlsx', data_only=True)
sheet = wb['Hoja 1']
rows = []
for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    if any(v is not None for v in vals):
        rows.append(vals)

print(f"Total rows in Excel: {len(rows)}")
header = rows[0]
print("Header:", header)
data_rows = rows[1:]

roles_map = {
    'ADMIN': 'ADMIN',
    'ADMINISTRADOR': 'ADMIN',
    'ADMINISTRADOR GENERAL': 'ADMIN',
    'MANAGEMENT': 'MANAGEMENT',
    'JEFATURA': 'MANAGEMENT',
    'JEFE': 'MANAGEMENT',
    'INVOICE_EXECUTOR': 'INVOICE_EXECUTOR',
    'EJECUTOR': 'INVOICE_EXECUTOR',
    'FACTURACIÓN': 'INVOICE_EXECUTOR',
    'FACTURACION': 'INVOICE_EXECUTOR',
    'EJECUTOR DE FACTURAS': 'INVOICE_EXECUTOR',
    'EJECUTOR DE FACTURACIÓN': 'INVOICE_EXECUTOR',
    'WAREHOUSE_USER': 'WAREHOUSE_USER',
    'SOLICITANTE': 'WAREHOUSE_USER',
    'BODEGA': 'WAREHOUSE_USER',
    'ENCARGADO DE BODEGA': 'WAREHOUSE_USER'
}

email_regex = re.compile(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$')

seen_emails = set()
processed = []
role_counts = {'WAREHOUSE_USER': 0, 'INVOICE_EXECUTOR': 0, 'MANAGEMENT': 0, 'ADMIN': 0, 'UNKNOWN': 0}
warehouses_count = {}
errors_list = []

for idx, r in enumerate(data_rows, start=2):
    nombre = str(r[0]).strip() if r[0] is not None else ''
    correo_raw = str(r[1]).strip() if r[1] is not None else ''
    bodega_raw = str(r[2]).strip() if r[2] is not None else ''
    rol_raw = str(r[3]).strip() if r[3] is not None else ''
    
    email_norm = correo_raw.lower()
    
    # Validation
    row_errors = []
    if not email_norm:
        row_errors.append('Email vacío')
    elif not email_regex.match(email_norm):
        row_errors.append(f'Email sintácticamente inválido: {email_norm}')
    elif email_norm in seen_emails:
        row_errors.append(f'Email duplicado en Excel: {email_norm}')
    else:
        seen_emails.add(email_norm)
        
    rol_norm = roles_map.get(rol_raw.upper())
    if not rol_norm:
        row_errors.append(f'Rol desconocido: {rol_raw}')
        role_counts['UNKNOWN'] += 1
    else:
        role_counts[rol_norm] += 1
        
    warehouse_candidate = ''
    if rol_norm == 'WAREHOUSE_USER':
        if nombre.lower().startswith('bodega '):
            warehouse_candidate = nombre[7:].strip()
        else:
            warehouse_candidate = nombre
        warehouses_count[warehouse_candidate] = warehouses_count.get(warehouse_candidate, 0) + 1
    else:
        warehouse_candidate = 'N/A'
            
    if row_errors:
        errors_list.append({'row': idx, 'nombre': nombre, 'email': email_norm, 'errors': row_errors})
        
    processed.append({
        'row': idx,
        'nombre': nombre,
        'email_raw': correo_raw,
        'email_norm': email_norm,
        'bodega_raw': bodega_raw,
        'rol_raw': rol_raw,
        'rol_norm': rol_norm,
        'warehouse_candidate': warehouse_candidate,
        'errors': row_errors
    })

print(f"\nTotal users in file: {len(processed)}")
print(f"Valid users: {len(processed) - len(errors_list)}")
print(f"Errors count: {len(errors_list)}")
print(f"Duplicate emails: {len(data_rows) - len(seen_emails)}")

print("\n--- ROLES COUNT ---")
for k, v in role_counts.items():
    print(f"{k}: {v}")

print("\n--- WAREHOUSE USERS BY BODEGA ---")
for k, v in warehouses_count.items():
    print(f"{k}: {v}")

print("\n--- ALL USERS TABLE ---")
print(f"{'Fila':<5} | {'Nombre':<24} | {'Email':<32} | {'Rol':<18} | {'Bodega':<18} | {'Resultado'}")
print("-" * 115)
for p in processed:
    res = "OK" if not p['errors'] else "ERROR: " + ", ".join(p['errors'])
    print(f"{p['row']:<5} | {p['nombre']:<24} | {p['email_norm']:<32} | {p['rol_norm']:<18} | {p['warehouse_candidate']:<18} | {res}")

if errors_list:
    print("\n--- ERRORS DETAIL ---")
    for e in errors_list:
        print(f"Row {e['row']}: {e['nombre']} ({e['email']}) -> {e['errors']}")
